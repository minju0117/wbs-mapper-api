"""
WBS 매핑 스크립트
다운로드된 WBS 데이터를 원본 WBS 템플릿에 매핑하여 출력합니다.

사용법:
  python3 wbs_mapper.py \
    --download /mnt/c/Users/uber/Downloads/download_WBS.xlsx \
    --template /mnt/c/Users/uber/Downloads/original_wbs.xlsx \
    --output /mnt/c/Users/uber/Downloads/output_WBS.xlsx \
    --sheet V1.4
"""

import argparse
import re
import shutil
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

# ── 템플릿 컬럼 위치 (1-based) ──────────────────────────────────────────────
COL_GUBUN    = 1   # A: 구분
COL_UPMU     = 2   # B: 업무
COL_DETAIL   = 3   # C: 세부 업무 항목
COL_OWNER    = 4   # D: 담당
COL_START    = 5   # E: 시작일
COL_END      = 6   # F: 종료일
COL_DAYS     = 7   # G: 기간 (수식 유지)
COL_DATE_PR  = 8   # H: 날짜기준 진척률 (수식 유지)
COL_ACTUAL   = 9   # I: 실제 진척률

TEMPLATE_DATA_START_ROW = 5   # 헤더 4행, 데이터 5행부터
HOLIDAY_REF  = "공휴일!$B$4:$B$15"
GANTT_DATE_ROW = 4    # 간트 날짜 헤더 행
GANTT_DATE_COL = 10   # J열: 간트 시작 날짜 (이후는 수식 자동계산)


def parse_date(val):
    """문자열 또는 datetime → datetime 변환"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(val), fmt)
        except ValueError:
            continue
    return val  # 변환 실패 시 원본 반환


def extract_version(path: str) -> str | None:
    """다운로드 WBS 상단 좌측에서 버전 문자열 추출 (예: v2, V1.4)"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in range(1, 6):
        val = ws.cell(row=row, column=1).value
        if val:
            m = re.search(r'[vV]\d[\d.]*', str(val))
            if m:
                return m.group(0)
    return None


def _is_version_sheet(name: str) -> bool:
    return bool(re.match(r'^[vV]\d', name.strip()))


def copy_previous_version_sheets(wb_target, prev_path: str, current_version: str | None):
    """이전 출력의 버전 시트들을 현재 workbook에 hidden으로 복사 (현재 버전 및 중복 제외)"""
    import copy
    wb_prev = openpyxl.load_workbook(prev_path)
    for name in wb_prev.sheetnames:
        if not _is_version_sheet(name):
            continue
        if name == current_version:
            continue
        if name in wb_target.sheetnames:
            continue
        ws_src = wb_prev[name]
        ws_dst = wb_target.create_sheet(title=name)
        ws_dst.sheet_state = "hidden"
        for row in ws_src.iter_rows():
            for cell in row:
                dst_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    dst_cell.font      = copy.copy(cell.font)
                    dst_cell.fill      = copy.copy(cell.fill)
                    dst_cell.border    = copy.copy(cell.border)
                    dst_cell.alignment = copy.copy(cell.alignment)
                    dst_cell.number_format = cell.number_format
        for col, cd in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col].width = cd.width
        for row_num, rd in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[row_num].height = rd.height
        for merged_range in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merged_range))


def load_download_rows(path: str) -> list:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # 컬럼 인덱스 동적 탐색 (헤더 이름 기준)
    idx = {name: i for i, name in enumerate(header) if name}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue
        rows.append({
            "구분":     row[idx.get("구분", 0)],
            "업무":     row[idx.get("업무", 1)],
            "세부업무":  row[idx.get("세부 업무 항목", 2)],
            "담당":     row[idx.get("담당", 3)],
            "시작일":   parse_date(row[idx.get("시작일", 4)]),
            "종료일":   parse_date(row[idx.get("종료일", 5)]),
            "진척률":   row[idx.get("진척률", 7)],  # '실제 진척률' 또는 '진척률'
        })
    return rows


def unmerge_ab_data_area(ws, data_start_row: int):
    """데이터 구역의 A·B열 병합 해제"""
    to_remove = [
        str(m) for m in ws.merged_cells.ranges
        if m.min_col <= 2 and m.min_row >= data_start_row
    ]
    for m in to_remove:
        ws.unmerge_cells(m)


def write_formula(ws, row: int, col: int, formula: str):
    ws.cell(row=row, column=col).value = formula


def fill_template(ws, dl_rows: list, data_start_row: int, gubun_fills: dict):
    template_last = ws.max_row

    for i, r in enumerate(dl_rows):
        out_row = data_start_row + i

        # 템플릿 행이 부족하면 마지막 행 스타일 복사로 추가
        if out_row > template_last:
            _copy_row_style(ws, template_last, out_row)

        ws.cell(row=out_row, column=COL_GUBUN).value   = r["구분"]  # 임시; 그룹화에서 정리
        ws.cell(row=out_row, column=COL_UPMU).value    = r["업무"]  # 임시; 그룹화에서 정리
        ws.cell(row=out_row, column=COL_DETAIL).value  = r["세부업무"]
        ws.cell(row=out_row, column=COL_OWNER).value   = r["담당"]
        ws.cell(row=out_row, column=COL_START).value   = r["시작일"]
        ws.cell(row=out_row, column=COL_END).value     = r["종료일"]
        ws.cell(row=out_row, column=COL_ACTUAL).value  = r["진척률"]

        # 수식: 기간, 날짜기준 진척률
        write_formula(ws, out_row, COL_DAYS,
            f"=NETWORKDAYS.INTL(E{out_row},F{out_row},1,{HOLIDAY_REF})")
        write_formula(ws, out_row, COL_DATE_PR,
            f'=IFERROR(MIN(1,(DATEDIF(E{out_row},TODAY(),"d")+1)'
            f'/(DATEDIF(E{out_row},F{out_row},"d")+1)),0)')

    # 간트 날짜 헤더(행4) + 월 레이블(행3) 업데이트
    start_dates = [r["시작일"] for r in dl_rows if isinstance(r["시작일"], datetime)]
    if start_dates:
        from datetime import timedelta
        earliest = min(start_dates)
        first_of_month = earliest.replace(day=1)
        gantt_start = first_of_month - timedelta(days=first_of_month.weekday())

        print(f"      간트 시작 날짜: {gantt_start.strftime('%Y-%m-%d')} "
              f"({['월','화','수','목','금','토','일'][gantt_start.weekday()]}요일)")

        # 행4: 수식 체인 대신 실제 날짜값으로 직접 채움 (캐시값 없는 수식 문제 방지)
        # col i 의 날짜 = gantt_start + (i + 2*(i//5)) days  (주말 2일씩 건너뜀)
        max_gantt_col = ws.max_column
        current_month_label_col = None
        prev_month = None

        for i in range(max_gantt_col - GANTT_DATE_COL + 1):
            col = GANTT_DATE_COL + i
            col_date = gantt_start + timedelta(days=i + 2 * (i // 5))
            cell = ws.cell(row=GANTT_DATE_ROW, column=col)
            cell.value = col_date

            # 행3 월 레이블: 새 월이 시작되면 레이블 업데이트
            month_val = col_date.month
            if month_val != prev_month:
                label_cell = ws.cell(row=GANTT_DATE_ROW - 1, column=col)
                label_cell.value = f"{month_val}월"
                if prev_month is not None and current_month_label_col is not None:
                    # 이전 월 레이블 셀 지우기 (이전 레이블이 이 위치에 있을 수 있음)
                    pass
                current_month_label_col = col
                prev_month = month_val

        # 기존 월 레이블 텍스트 정리: 새로 쓴 위치 외 레이블 삭제
        written_cols = set()
        prev_m = None
        for i in range(max_gantt_col - GANTT_DATE_COL + 1):
            col = GANTT_DATE_COL + i
            col_date = gantt_start + timedelta(days=i + 2 * (i // 5))
            if col_date.month != prev_m:
                written_cols.add(col)
                prev_m = col_date.month

        for col in range(GANTT_DATE_COL, max_gantt_col + 1):
            cell = ws.cell(row=GANTT_DATE_ROW - 1, column=col)
            if cell.value is not None and col not in written_cols:
                cell.value = None

    # A/B 열 그룹화: 같은 값이 연속되면 첫 행만 값 유지, B열은 병합
    _apply_grouping(ws, dl_rows, data_start_row)
    # 그룹 경계 보더 정렬
    _fix_group_borders(ws, dl_rows, data_start_row)
    # 그룹별 색상 적용
    _apply_group_fills(ws, dl_rows, data_start_row, gubun_fills)

    # 초과 템플릿 행: E·F 초기화 (간트 막대 제거), 나머지 값도 클리어
    last_written = data_start_row + len(dl_rows) - 1
    for row in range(last_written + 1, template_last + 1):
        for col in range(1, COL_ACTUAL + 1):
            ws.cell(row=row, column=col).value = None


def _apply_grouping(ws, dl_rows: list, data_start_row: int):
    """
    A열(구분): 같은 구분이 연속될 때 첫 행만 값 유지, 나머지 None
    B열(업무): 같은 업무가 연속될 때 첫 행만 값 유지 + 해당 구간 셀 병합
    """
    import copy

    n = len(dl_rows)

    # ── A열: 구분 그룹 첫 행만 표시 ──────────────────────────────────────
    prev_gubun = None
    for i, r in enumerate(dl_rows):
        row = data_start_row + i
        if r["구분"] == prev_gubun:
            ws.cell(row=row, column=COL_GUBUN).value = None
        else:
            prev_gubun = r["구분"]

    # ── B열: 업무 그룹 병합 ──────────────────────────────────────────────
    # 연속된 (구분, 업무) 동일 구간을 찾아 병합
    prev_key = None
    group_start = data_start_row

    def _merge_b_group(start_row, end_row):
        if start_row >= end_row:
            # 단일 행: 병합 불필요, 값만 첫 행 유지
            return
        # 첫 행 스타일 가져오기
        master = ws.cell(row=start_row, column=COL_UPMU)
        master_style = {
            "font":      copy.copy(master.font),
            "fill":      copy.copy(master.fill),
            "border":    copy.copy(master.border),
            "alignment": copy.copy(master.alignment),
            "number_format": master.number_format,
        }
        # 나머지 행 값 클리어 후 스타일 복사
        for r in range(start_row + 1, end_row + 1):
            cell = ws.cell(row=r, column=COL_UPMU)
            cell.value = None
            cell.font      = master_style["font"]
            cell.fill      = master_style["fill"]
            cell.border    = master_style["border"]
            cell.alignment = master_style["alignment"]
            cell.number_format = master_style["number_format"]
        # 셀 병합
        ws.merge_cells(
            start_row=start_row, start_column=COL_UPMU,
            end_row=end_row,     end_column=COL_UPMU
        )

    for i, r in enumerate(dl_rows):
        row = data_start_row + i
        key = (r["구분"], r["업무"])
        if key != prev_key:
            if prev_key is not None:
                _merge_b_group(group_start, row - 1)
            group_start = row
            prev_key = key
        else:
            ws.cell(row=row, column=COL_UPMU).value = None

    # 마지막 그룹 처리
    if prev_key is not None:
        _merge_b_group(group_start, data_start_row + n - 1)


def _apply_group_fills(ws, dl_rows: list, data_start_row: int, gubun_fills: dict):
    """구분 그룹별 A열 배경색 적용. B~I열은 공통 연회색."""
    import copy
    from openpyxl.styles import PatternFill

    # B~I 공통 연회색
    gray_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    default_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    # 구분 이름 정규화 (공백 제거, 소문자)
    normalized = {k.replace(" ", "").lower(): v for k, v in gubun_fills.items()}

    for i, r in enumerate(dl_rows):
        row = data_start_row + i
        gubun = (r["구분"] or "").replace(" ", "").lower()
        group_fill = normalized.get(gubun, default_fill)

        ws.cell(row=row, column=COL_GUBUN).fill = copy.copy(group_fill)
        for col in range(COL_UPMU, COL_ACTUAL + 1):
            ws.cell(row=row, column=col).fill = copy.copy(gray_fill)


def _fix_group_borders(ws, dl_rows: list, data_start_row: int):
    """그룹 경계 행에 top/bottom 보더를 새 위치에 맞게 재설정."""
    from openpyxl.styles import Border, Side
    import copy

    thin     = Side(border_style="thin")
    no_side  = Side(border_style=None)

    n = len(dl_rows)
    max_col = ws.max_column

    # 그룹 시작/끝 행 번호 수집
    group_starts: set[int] = set()
    group_ends:   set[int] = set()
    prev_gubun = None
    for i, r in enumerate(dl_rows):
        row = data_start_row + i
        if r["구분"] != prev_gubun:
            if prev_gubun is not None:
                group_ends.add(row - 1)
            group_starts.add(row)
            prev_gubun = r["구분"]
    group_ends.add(data_start_row + n - 1)

    # bottom 적용 컬럼: A B C D F G H I (E 제외)
    BOTTOM_COLS = {COL_GUBUN, COL_UPMU, COL_DETAIL, COL_OWNER,
                   COL_END, COL_DAYS, COL_DATE_PR, COL_ACTUAL}
    # top 적용 컬럼: A E (첫 그룹 제외)
    TOP_COLS = {COL_GUBUN, COL_START}
    first_row = data_start_row

    for i in range(n):
        row = data_start_row + i
        is_start = row in group_starts and row != first_row  # 첫 그룹엔 top 없음
        is_end   = row in group_ends

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            b    = cell.border

            if col > COL_ACTUAL:
                # 간트 차트 영역: 그룹 끝 행에만 bottom 선 적용
                want_top    = no_side
                want_bottom = thin if is_end else no_side
            else:
                want_top    = thin if (is_start and col in TOP_COLS)    else no_side
                want_bottom = thin if (is_end   and col in BOTTOM_COLS) else no_side

            # 기존 left/right 유지하면서 top/bottom만 교체
            cell.border = Border(
                top    = want_top,
                bottom = want_bottom,
                left   = copy.copy(b.left)  if b and b.left  else no_side,
                right  = copy.copy(b.right) if b and b.right else no_side,
            )


def _copy_row_style(ws, src_row: int, dst_row: int):
    """스타일 참조용: 마지막 데이터 행의 셀 스타일을 신규 행에 복사"""
    import copy
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font      = copy.copy(src_cell.font)
            dst_cell.fill      = copy.copy(src_cell.fill)
            dst_cell.border    = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format


def main():
    parser = argparse.ArgumentParser(description="WBS 매핑 스크립트")
    parser.add_argument("--download",  required=True, help="다운로드된 WBS 파일 경로")
    parser.add_argument("--template",  required=True, help="원본 WBS 템플릿 파일 경로")
    parser.add_argument("--output",    required=True, help="출력 파일 경로")
    parser.add_argument("--sheet",     default="V1.4", help="사용할 템플릿 시트 이름 (기본: V1.4)")
    parser.add_argument("--previous",  default=None,  help="이전 출력 WBS 파일 경로 (버전 히스토리 포함)")
    args = parser.parse_args()

    print(f"[1/4] 다운로드 WBS 읽기: {args.download}")
    dl_rows = load_download_rows(args.download)
    print(f"      → {len(dl_rows)}개 행 로드")

    # 원본 템플릿에서 구분별 A열 fill 색상 수집
    import copy as _copy
    wb_orig = openpyxl.load_workbook(args.template)
    ws_orig = wb_orig[args.sheet if args.sheet in wb_orig.sheetnames else wb_orig.sheetnames[0]]
    gubun_fills = {}
    for row in range(TEMPLATE_DATA_START_ROW, ws_orig.max_row + 1):
        val = ws_orig.cell(row, COL_GUBUN).value
        if val and val not in gubun_fills:
            gubun_fills[val] = _copy.copy(ws_orig.cell(row, COL_GUBUN).fill)
    print(f"      구분 색상 수집: {list(gubun_fills.keys())}")

    print(f"[2/4] 템플릿 복사: {args.template} → {args.output}")
    shutil.copy2(args.template, args.output)

    print(f"[3/4] 데이터 매핑 (시트: {args.sheet})")
    wb = openpyxl.load_workbook(args.output)

    # 이전 출력의 버전 시트 복사 (히스토리 누적)
    version = extract_version(args.download)
    if args.previous:
        print(f"      이전 출력에서 버전 시트 복사: {args.previous}")
        copy_previous_version_sheets(wb, args.previous, version)

    # 히든 시트 삭제 (공휴일·버전 패턴 시트는 보존)
    hidden_sheets = [
        name for name in wb.sheetnames
        if wb[name].sheet_state in ("hidden", "veryHidden")
        and name != "공휴일"
        and not _is_version_sheet(name)
    ]
    for name in hidden_sheets:
        del wb[name]
    if hidden_sheets:
        print(f"      히든 시트 삭제: {hidden_sheets}")

    if args.sheet not in wb.sheetnames:
        print(f"      경고: '{args.sheet}' 시트 없음. 사용 가능: {wb.sheetnames}")
        sheet_name = wb.sheetnames[0]
    else:
        sheet_name = args.sheet

    ws = wb[sheet_name]

    # 버전 시트 생성: 다운로드 WBS 상단에서 버전 읽어 새 시트로 복사, 기존 시트는 hidden
    if version and version != sheet_name:
        print(f"      버전 시트 생성: {sheet_name} → {version}")
        new_ws = wb.copy_worksheet(ws)
        new_ws.title = version
        ws.sheet_state = "hidden"
        ws = new_ws

    unmerge_ab_data_area(ws, TEMPLATE_DATA_START_ROW)
    fill_template(ws, dl_rows, TEMPLATE_DATA_START_ROW, gubun_fills)

    # 데이터 이후 빈 행 삭제
    last_data_row = TEMPLATE_DATA_START_ROW + len(dl_rows) - 1
    if ws.max_row > last_data_row:
        rows_to_delete = ws.max_row - last_data_row
        ws.delete_rows(last_data_row + 1, rows_to_delete)
        print(f"      빈 행 {rows_to_delete}개 삭제 (행{last_data_row+1}~)")

    print(f"[4/4] 저장: {args.output}")
    wb.save(args.output)
    print("완료!")


if __name__ == "__main__":
    main()
