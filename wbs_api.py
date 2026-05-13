"""
WBS 매퍼 FastAPI 서버
POST /map-wbs  — download WBS 업로드 → 원본 템플릿과 매핑 → 완성된 엑셀 반환
템플릿 파일은 서버에 original_wbs.xlsx 로 번들되어 있음
"""

import copy
import io
import json
import os
import shutil
import tempfile
from typing import List, Optional

import anthropic
import openpyxl
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from wbs_mapper import (
    COL_GUBUN,
    TEMPLATE_DATA_START_ROW,
    copy_previous_version_sheets,
    extract_header_info,
    fill_template,
    load_download_rows,
    unmerge_ab_data_area,
    update_holiday_sheet,
)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "original_wbs.xlsx")

app = FastAPI(title="WBS Mapper API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["POST", "GET"],
    allow_headers=["*"],
)


class MenuInfo(BaseModel):
    name: str
    phase: str = "1차"
    pages: int = 1


class ScheduleRequest(BaseModel):
    project: str
    client: str
    start_date: str
    first_launch_date: Optional[str] = None
    end_date: str
    multilingual: bool = False
    conditions: Optional[str] = None
    menus: List[MenuInfo] = []


class AdjustRequest(BaseModel):
    project: str
    client: str
    start_date: str
    end_date: str
    first_launch_date: Optional[str] = None
    conditions: Optional[str] = None
    tasks: List[dict] = []


SCHEDULE_SYSTEM_PROMPT = """[출력 규칙] 반드시 JSON 배열([...])만 출력. 분석, 설명, 주석, 마크다운 코드블록 절대 금지. 첫 글자는 반드시 [, 마지막 글자는 반드시 ].

당신은 한국 웹 에이전시 "더위버"의 WBS 일정 전문가입니다.
6개 실제 프로젝트(DB그룹, DB하이텍, 효성그룹, 대상웰라이프, 삼양그룹, 삼화페인트)를 분석한 패턴 기반으로 일정을 생성합니다.

## 담당 표기
- 더위버 단독 작업: "더위버"
- 고객사 단독 작업: "{입력된 고객사명}" (콘텐츠수급, 보안점검 등)
- 협업: "{입력된 고객사명}/더위버" (킥오프, 요구사항정의, 서버협의, 검수, 오픈 등)
- 다국어 번역: "더위버/번역업체"

## 날짜 계산 순서 (반드시 이 순서로 계산)

### STEP 1. 착수 단계
start_date를 D0으로 한다.

메뉴 수 기준: 소규모 = 메뉴 1~5개, 대규모 = 메뉴 6개 이상

| 업무 | task | subTask | startDate | duration(소규모) | duration(대규모) | owner |
|------|------|---------|-----------|----------------|----------------|-------|
| 착수 | 기획 | 킥오프 | D0 | 1 | 1 | {고객사명}/더위버 |
| 착수 | 기획 | 구축방향/요구사항정의 | D0 | 2~4 | 4~9 | {고객사명}/더위버 |
| 착수 | 기획 | WBS수립 | D0+1 | 3~4 | 6~8 | 더위버 |
| 착수 | 기획 | IA설계 | D0+1 | 3~5 | 7~10 | 더위버 | ← WBS수립과 병행, WBS보다 1~2일 늦게 끝남

IA_END = D0+1 + (IA설계 기간) - 1  ← 이후 모든 Front 화면설계의 기준점

| 착수 | 기획 | 콘텐츠 수급 - {메뉴명} | D0+1 | 5~15 | 10~20 | {고객사명} | ← 메뉴마다 1행씩, 반드시 전체 메뉴 포함
| 착수 | 개발 | 서버 환경 확인 및 설정 협의 | IA_END | 4 | 5 | {고객사명}/더위버 |
| 착수 | 개발 | 내부 개발 환경 세팅 | 서버협의 완료 후 | 4 | 5 | 더위버 |

### STEP 2. Admin 단계 (IA_END 이후 시작)
| Admin | 기획 | 관리자 기능 정의 | IA_END | 4~13 | 더위버 |
| Admin | 기획 | 화면설계 | 관리자기능정의 완료 후 | 5~19 | 더위버 |
| Admin | 개발 | 설계/협의 | Admin 화면설계 시작과 동시 | 5 | 더위버 |
| Admin | 개발 | 관리자 개발 | Admin 화면설계 완료 후 | 14~31 | 더위버 |

### STEP 3. Front 화면설계 (IA_END 다음 영업일부터 시작, Admin과 병행)
⚠️ 화면설계는 절대 IA_END 이전에 시작하지 않는다.

메뉴 순서대로 순차 진행 (1차 오픈 메뉴 우선):
| Front | 기획 | 화면설계 - 메인 | IA_END | 2~7 | 더위버 |
| Front | 기획 | 화면설계 - {메뉴1} | 메인 완료 후 | 2~6 | 더위버 |
| Front | 기획 | 화면설계 - {메뉴2} | 메뉴1 완료 후 | 2~6 | 더위버 |
(메뉴 수가 많으면 2~3개 병행 가능)

### STEP 4. 실행안/시안 제작 (IA_END 즈음 시작, 화면설계와 병행)
| Front | 디자인 | 실행안 제작 | IA_END | 4~6 | 더위버 |
| Front | 디자인 | 스타일 가이드 제작 | 실행안 완료 후 | 2~6 | 더위버 |

### STEP 5. 페이지 디자인 (각 메뉴 화면설계 완료 후 시작)
⚠️ 각 메뉴의 디자인은 반드시 해당 메뉴 화면설계 완료 후 시작.
| Front | 디자인 | 페이지 디자인 - 메인 | 메인 화면설계 완료 후 | 3~6 | 더위버 |
| Front | 디자인 | 페이지 디자인 - {메뉴} | 해당 메뉴 화면설계 완료 후 | 2~14 | 더위버 |

### STEP 6. 퍼블리싱 기반 작업 (첫 메뉴 디자인 완료 즈음 시작)
| Front | 퍼블리싱 | 반응형웹 제작 기준 수립 | 첫 디자인 완료 즈음 | 4~5 | 더위버 |
| Front | 퍼블리싱 | 개발환경 세팅, 코딩 가이드 제작 | 반응형 기준 완료 후 | 2~10 | 더위버 |

### STEP 7. 페이지 퍼블리싱 (각 메뉴 디자인 완료 후 시작)
⚠️ 각 메뉴의 퍼블리싱은 반드시 해당 메뉴 디자인 완료 후 시작.
| Front | 퍼블리싱 | 페이지 퍼블리싱 - 메인 | 메인 디자인 완료 후 | 3~10 | 더위버 |
| Front | 퍼블리싱 | 페이지 퍼블리싱 - {메뉴} | 해당 메뉴 디자인 완료 후 | 2~10 | 더위버 |

⚠️ 퍼블리싱 마감 — 역산으로 계산 (절대 규칙):
오픈일에서 거꾸로 계산하여 각 퍼블리싱의 완료일이 오픈일을 넘지 않도록 duration을 결정한다.
- 단일 오픈: 마지막 퍼블리싱 완료 ≤ 운영배포일 - (검수 3~5일 + 개발 12~20일 + 기능검수 4~10일 + 고객사검수 5~13일 + 통합테스트 3~5일 + 여유 2일)
- 1차 오픈: 1차 메뉴 마지막 퍼블 완료 ≤ first_launch_date - (1차퍼블검수 잔여일 + 통합테스트 3일 + 서버포팅 1일)
- 2차 오픈: 2차 메뉴 마지막 퍼블 완료 ≤ end_date - (2차퍼블검수 잔여일 + 2차개발 5일 + 서버포팅 1일)
- 기간이 빠듯하면 각 메뉴 퍼블 duration을 최솟값(2~3일)으로 줄인다.

### STEP 8. 퍼블리싱 검수 및 수정

⚠️ 검수 시작/종료 규칙 (실제 프로젝트 패턴 기반):
- 검수 시작 = 해당 차수 전체 메뉴의 절반 이상 퍼블 완료 시점
- 검수 종료 = 해당 차수 마지막 메뉴 퍼블 완료일 + 3~5영업일
  (퍼블리싱이 모두 끝난 후에도 3~5일 더 이어진다)
- 실제 예시: 효성그룹 마지막퍼블 04/11 → 검수 04/07~04/15(+4일), 대상웰라이프 마지막퍼블 02/21 → 검수 02/10~02/28(+5일)

| Front | 내부검수 | 퍼블리싱 검수 및 수정 | 절반 이상 퍼블 완료 후 | 마지막퍼블+3~5일까지 | 더위버 |
| Front | 내부검수 | 중간보고 | 검수 중 | 1 | {고객사명}/더위버 |
| Front | 개발 | 개발 | 퍼블리싱 검수 완료 후 | 12~20 | 더위버 |
| Front | 검수 | 기능 검수 및 수정 | 개발 완료 후 | 4~10 | 더위버 |
| Front | 검수 | 고객사 검수 및 수정 | 기능검수 완료 후 | 5~13 | {고객사명} |

### STEP 9. 오픈

⚠️ 역산 원칙: 반드시 오픈일에서 역산하여 일정 배치. 절대로 오픈일을 초과하지 않는다.

**단일 오픈 (first_launch_date 없음):**
운영배포일(end_date) ← 통합테스트(3~5일) ← 고객사검수 ← 기능검수 ← 개발 ← 검수 ← 퍼블
| 테스트/오픈 | 통합테스트 | 통합 테스트 및 검수 | 고객사검수 완료 후 | 3~5 | {고객사명}/더위버 |
| 테스트/오픈 | 오픈 | 운영 배포 | 통합테스트 완료 후 | 1 | {고객사명}/더위버 |
| 산출물작성 | (빈값) | (빈값) | 오픈 후 | 5~9 | 더위버 |

**1차/2차 오픈 (first_launch_date 있음):**
[역산] first_launch_date ← 1차서버포팅(1일) ← 통합테스트(3일) ← 1차퍼블검수 종료
→ 1차퍼블검수 종료 = 1차 마지막 퍼블 완료 + 3~5일
→ 1차 마지막 퍼블 완료 = first_launch_date - 서버포팅(1) - 통합테스트(3) - 검수잔여(3~5)

[역산] end_date ← 2차서버포팅(1일) ← 2차개발(5일) ← 2차퍼블검수 종료
→ 2차퍼블검수 종료 = 2차 마지막 퍼블 완료 + 3~5일
→ 2차 마지막 퍼블 완료 = end_date - 서버포팅(1) - 2차개발(5) - 검수잔여(3~5)

산출물작성 (2차 오픈 후, 5일)

### 다국어 지원 시 (multilingual=true, 국문 퍼블 완료 후 추가)
| 번역 | 번역 | 영·중문 번역 | 국문퍼블 완료 후 | 15 | 더위버/번역업체 |
| 번역 | 번역 | 고객사 내부 번역 검수 | 번역 완료 후 | 9 | {고객사명} |
| Front(영,중문) | 기획 | 화면설계 | 번역 완료 즈음 | 5 | 더위버 |
| Front(영,중문) | 디자인 | 디자인 가이드 | 화면설계 후 | 5 | 더위버 |
| Front(영,중문) | 퍼블리싱 | 퍼블리싱 | 디자인가이드 후 | 15 | 더위버 |
| Front(영,중문) | 개발 | 개발 | 퍼블리싱 후 | 5 | 더위버 |

## ⚠️ 마감일 절대 준수
- 생성되는 모든 태스크의 endDate는 end_date를 초과할 수 없다 (산출물작성 포함).
- first_launch_date가 있을 때, 1차 관련 태스크(1차 오픈까지)의 endDate는 first_launch_date를 초과할 수 없다.
- 일정이 촉박하면 각 태스크 duration을 최솟값으로 줄여서라도 마감일 내에 완료한다.
- 절대로 오픈일 이후에 퍼블리싱 태스크가 배치되면 안 된다.

## 출력 형식
JSON 배열만 출력. 다른 텍스트 없이. 7개 필드만 포함:
[{"category":"착수","task":"기획","subTask":"킥오프","owner":"{고객사명}/더위버","startDate":"YYYY-MM-DD","endDate":"YYYY-MM-DD","duration":1}]
공백 최소화."""


ADJUST_SYSTEM_PROMPT = """[출력 규칙] 반드시 JSON 배열([...])만 출력. 분석·설명·주석·코드블록 절대 금지. 첫 글자 [, 마지막 글자 ].

당신은 한국 웹 에이전시 "더위버"의 WBS 일정 조정 전문가입니다.

## 작업 절차

### STEP 1. locked=true 보존
locked=true 태스크의 category/task/subTask/owner/startDate/endDate/duration 모두 변경 금지.

### STEP 2. 사용자 요청사항(conditions) 반영
conditions에 명시된 내용을 반영하여 해당 태스크의 startDate/endDate/duration을 조정한다.

### STEP 3. 오픈일 역산으로 전체 일정 재검토
STEP 2 적용 결과가 오픈일을 초과하는지 반드시 검토한다.
초과하면 다음 방법으로 오픈일 내에 맞춘다:
- 각 퍼블리싱 태스크의 duration을 줄인다 (최솟값 2일)
- 퍼블리싱을 앞당겨 배치한다 (단, 해당 메뉴 디자인 완료 후여야 함)
- 빠듯하더라도 오픈일을 절대 넘기지 않는다

오픈일 기준:
- first_launch_date가 있으면: 1차 관련 태스크 전체 endDate ≤ first_launch_date
- end_date: 모든 태스크 endDate ≤ end_date
- 퍼블리싱은 오픈일 이전에 완료. 오픈 당일 이후 배치 금지.

의존관계 (전체 일정 재검토 시 준수):
- 화면설계 → IA설계 완료 후
- 페이지 디자인 → 해당 메뉴 화면설계 완료 후
- 퍼블리싱 → 해당 메뉴 디자인 완료 후
- 개발 → 퍼블리싱 검수 완료 후
- 고객사 검수 → 기능검수 완료 후
- 통합테스트 → 고객사 검수 완료 후

### STEP 4. 검수 기간 보정
퍼블리싱 검수 태스크(subTask에 "검수" 포함)의 endDate를 확인한다.
해당 차수 마지막 페이지 퍼블리싱 endDate + 2영업일 미만이면 연장한다.
단, 연장 후에도 오픈일(first_launch_date 또는 end_date)을 초과하면 안 된다.

### 공통 제약
- 입력 tasks 배열에 없는 태스크 추가 금지. 삭제된 항목 복원 금지. 태스크 수·순서 유지.
- 주말(토/일) 제외. duration은 영업일 기준.

## 출력 형식
입력 tasks 전체 반환. locked=true도 원본 그대로 포함.
7개 필드: category, task, subTask, owner, startDate(YYYY-MM-DD), endDate(YYYY-MM-DD), duration
JSON 배열만. 설명 없이. 공백 최소화. [로 시작 ]로 끝."""


@app.get("/")
def health():
    return {"status": "ok"}


@app.post("/adjust-schedule")
async def adjust_schedule(request: AdjustRequest):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY 환경변수가 설정되지 않았습니다.")

    if not request.tasks:
        raise HTTPException(status_code=400, detail="tasks 배열이 비어 있습니다.")

    locked_count = sum(1 for t in request.tasks if t.get("locked"))

    # Claude에게는 일정 관련 7개 필드 + locked만 전달 (토큰 절약)
    KEEP_FIELDS = {"category", "task", "subTask", "owner", "startDate", "endDate", "duration", "locked"}
    slim_tasks = [{k: v for k, v in t.items() if k in KEEP_FIELDS} for t in request.tasks]
    tasks_json = json.dumps(slim_tasks, ensure_ascii=False, separators=(",", ":"))

    # 원본 extra 필드 보존 (인덱스 기준)
    extra_fields_by_index = [
        {k: v for k, v in t.items() if k not in KEEP_FIELDS}
        for t in request.tasks
    ]

    conditions_block = f"""
## ★ 사용자 요청사항 (최우선 반영)
{request.conditions}
""" if request.conditions else ""

    user_message = f"""다음 WBS 일정을 조정해주세요.

프로젝트명: {request.project}
고객사: {request.client}
시작일: {request.start_date}
종료일(절대 초과 금지): {request.end_date}
1차 오픈일(절대 초과 금지): {request.first_launch_date or '없음'}
{conditions_block}
현재 tasks — 아래 배열이 전부다. 이 외의 태스크를 추가하거나 삭제된 항목을 복원하지 말 것.
locked=true({locked_count}개)는 날짜/내용 변경 불가. locked=false는 조정 대상.
{tasks_json}

반드시 JSON 배열([...])만 출력. [ 로 시작하는 JSON만. 설명 없이."""

    try:
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=16000,
            system=ADJUST_SYSTEM_PROMPT,
            messages=[
                {"role": "user", "content": user_message},
            ],
        )
        raw = message.content[0].text.strip()

        if "```" in raw:
            parts = raw.split("```")
            for part in parts:
                part = part.strip()
                if part.startswith("json"):
                    part = part[4:].strip()
                if part.startswith("["):
                    raw = part
                    break

        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1 and end > start:
            raw = raw[start:end + 1]

        tasks = json.loads(raw)

        # extra 필드 복원 + 기본값 설정
        for i, t in enumerate(tasks):
            if i < len(extra_fields_by_index):
                for k, v in extra_fields_by_index[i].items():
                    t.setdefault(k, v)
            t.setdefault("scheduledProgress", 0)
            t.setdefault("actualProgress", 0)
            t.setdefault("progress", 0)
            t.setdefault("locked", False)
            t.setdefault("isCompleted", False)
            t.setdefault("delayReason", "")

    except json.JSONDecodeError as e:
        preview = raw[:300] if 'raw' in dir() else "N/A"
        raise HTTPException(status_code=500, detail=f"Claude 응답 파싱 오류: {e} | 응답 앞부분: {preview}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API 오류: {e}")

    return {"tasks": tasks}


@app.post("/generate-schedule")
async def generate_schedule(request: ScheduleRequest):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY 환경변수가 설정되지 않았습니다.")

    menu_lines = "\n".join(
        f"  - {m.name}: {m.phase} 오픈, {m.pages}페이지"
        for m in request.menus
    )

    user_message = f"""다음 웹 프로젝트의 WBS 일정을 생성해주세요.

프로젝트명: {request.project}
고객사: {request.client}
시작일: {request.start_date}
1차 오픈일: {request.first_launch_date or '없음'}
종료일: {request.end_date}
다국어 지원: {'예' if request.multilingual else '아니오'}
기타 조건: {request.conditions or '없음'}

메뉴 구성:
{menu_lines}

위 정보를 바탕으로 업무 의존관계를 지키는 현실적인 WBS 일정을 생성해주세요.

반드시 JSON 배열([...])만 출력. 분석이나 설명 없이 [ 로 시작하는 JSON만."""

    try:
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=16000,
            system=SCHEDULE_SYSTEM_PROMPT,
            messages=[
                {"role": "user", "content": user_message},
            ],
        )
        raw = message.content[0].text.strip()

        # ``` 블록 제거
        if "```" in raw:
            parts = raw.split("```")
            for part in parts:
                part = part.strip()
                if part.startswith("json"):
                    part = part[4:].strip()
                if part.startswith("["):
                    raw = part
                    break

        # [ ... ] 배열 직접 추출 (앞뒤 설명 텍스트 제거)
        start = raw.find("[")
        end = raw.rfind("]")
        if start != -1 and end != -1 and end > start:
            raw = raw[start:end + 1]

        tasks = json.loads(raw)

        # 기본값 필드 자동 주입
        for t in tasks:
            t.setdefault("scheduledProgress", 0)
            t.setdefault("actualProgress", 0)
            t.setdefault("progress", 0)
            t.setdefault("locked", False)
            t.setdefault("isCompleted", False)
            t.setdefault("delayReason", "")

    except json.JSONDecodeError as e:
        preview = raw[:300] if 'raw' in dir() else "N/A"
        raise HTTPException(status_code=500, detail=f"Claude 응답 파싱 오류: {e} | 응답 앞부분: {preview}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API 오류: {e}")

    return {"tasks": tasks}


@app.post("/map-wbs")
async def map_wbs(
    download: UploadFile = File(..., description="Lovable에서 다운받은 WBS xlsx"),
    previous: UploadFile | None = File(None, description="이전 출력 WBS (버전 히스토리용, 선택)"),
    version: str | None = Query(None, description="WBS 버전 (예: v2)"),
    project: str | None = Query(None, description="프로젝트명"),
    sheet: str = Query(default="V1.4", description="템플릿 시트 이름"),
):
    if not download.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=f"{download.filename}: xlsx 파일만 허용됩니다.")

    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail="서버에 원본 템플릿 파일이 없습니다.")

    with tempfile.TemporaryDirectory() as tmpdir:
        dl_path  = os.path.join(tmpdir, "download.xlsx")
        tpl_path = os.path.join(tmpdir, "template.xlsx")
        out_path = os.path.join(tmpdir, "output.xlsx")

        with open(dl_path, "wb") as f:
            f.write(await download.read())

        prev_path = None
        if previous:
            prev_path = os.path.join(tmpdir, "previous.xlsx")
            with open(prev_path, "wb") as f:
                f.write(await previous.read())

        shutil.copy2(TEMPLATE_PATH, tpl_path)

        try:
            dl_rows = load_download_rows(dl_path)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"다운로드 WBS 파싱 오류: {e}")

        # query param 미전달 시 다운로드 WBS 상단에서 자동 추출
        if not project or not version:
            extracted_project, extracted_version = extract_header_info(dl_path)
            if not project:
                project = extracted_project
            if not version:
                version = extracted_version

        if not dl_rows:
            raise HTTPException(status_code=422, detail="다운로드 WBS에 데이터가 없습니다.")

        # 템플릿에서 구분별 fill 색상 수집
        wb_orig = openpyxl.load_workbook(tpl_path)
        sheet_name = sheet if sheet in wb_orig.sheetnames else wb_orig.sheetnames[0]
        ws_orig = wb_orig[sheet_name]
        gubun_fills = {}
        for row in range(TEMPLATE_DATA_START_ROW, ws_orig.max_row + 1):
            val = ws_orig.cell(row, COL_GUBUN).value
            if val and val not in gubun_fills:
                gubun_fills[val] = copy.copy(ws_orig.cell(row, COL_GUBUN).fill)

        shutil.copy2(tpl_path, out_path)
        wb = openpyxl.load_workbook(out_path)

        # ① 히든 시트 전부 삭제 (공휴일만 보존)
        for name in [
            n for n in wb.sheetnames
            if wb[n].sheet_state in ("hidden", "veryHidden") and n != "공휴일"
        ]:
            del wb[name]

        # ② 이전 출력의 버전 시트를 hidden으로 복사 (삭제 이후에 추가)
        if prev_path:
            copy_previous_version_sheets(wb, prev_path, version)

        # ③ 작업 시트 선택 및 시트명 버전으로 변경
        ws = wb[sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]]
        if version:
            ws.title = version

        # ④ 프로젝트명(A1), Last update(A2) 갱신
        from datetime import datetime as _dt
        if project:
            ws.cell(row=1, column=1).value = project
        ws.cell(row=2, column=1).value = f"▥ Last update : {_dt.today().strftime('%Y-%m-%d')}"

        # 공휴일 시트 자동 업데이트
        from datetime import datetime as _dt2
        dates = [r[k] for r in dl_rows for k in ("시작일", "종료일") if isinstance(r.get(k), _dt2)]
        if dates:
            sy, ey = min(d.year for d in dates), max(d.year for d in dates)
        else:
            sy = ey = _dt2.today().year
        holiday_ref = update_holiday_sheet(wb, sy, ey)

        unmerge_ab_data_area(ws, TEMPLATE_DATA_START_ROW)
        fill_template(ws, dl_rows, TEMPLATE_DATA_START_ROW, gubun_fills, holiday_ref)

        last_data_row = TEMPLATE_DATA_START_ROW + len(dl_rows) - 1
        if ws.max_row > last_data_row:
            ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)

        wb.save(out_path)

        with open(out_path, "rb") as f:
            content = f.read()

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="output_WBS.xlsx"'},
    )
